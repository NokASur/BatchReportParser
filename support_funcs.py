import re

from openpyxl.utils import get_column_letter


def concatenate_list_values(l: list) -> str:
    return ", ".join(str(v) for v in l)


def remove_extra_spaces(input_string):
    return re.sub(r'\s{2,}', ' ', input_string)


# writes meta data to the requires sheet in a reasonable way
def imprint_meta_values_to_sheet(sheet, meta_values: list[list[int | str | float]], y: int = 0, x: int = 0):
    for values in meta_values:
        bias = 0
        for value in values:
            sheet.cell(row=y, column=x + bias).value = value
            bias += 1
        y += 1


# adjusts sheet cells to be of the correct length
def adjust_excel_cells_length(writer, sheet_name: str):
    sheet = writer.sheets[sheet_name]
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column
        column_letter = get_column_letter(column)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception as e:
                print(f"Mistake while adjusting cell: {e}")
        sheet.column_dimensions[column_letter].width = max_length + 5


# for now just checks if there are only "кызыл", "армян" named batches in mistakes,
# as they are deemed unimportant and should not be punished
def only_unimportant_mistakes(mistakes: list[str]) -> bool:
    unimportant_names = ["кызыл", "армян"]
    for mistake in mistakes:
        good = 0
        mistake_name = mistake.lower()
        for unimportant_name in unimportant_names:
            if unimportant_name in mistake_name:
                good = 1
        if good == 0:
            return False
    return True
